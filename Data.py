from openpyxl import load_workbook


class DB:
    def __init__(self, location:str, sheet:str='Sheet1'):
        self.wb = load_workbook(location)
        self.ws = self.wb[sheet]
        self.location = location
        self.head = [x.lower() for x in self.GetHead()]
    
    def Loadin(self, data:dict, ntype=dict, defin=0):
        mis = []
        new = []
        for i in self.head:
            if i in data:
                new.append(data[i])
            else:
                mis.append(i)
                new.append(defin)
        if ntype == dict:
            new = dict(zip(self.head,new))
            return new, mis
        elif ntype == list:
            return new
    
    def LoadinExt(self, data:dict, row:int, stop:list=[]):
        org = self.GetRows([row])
        if not org:return

        org = org[0]
        new = []
        for i in self.head:
            if i in stop:
                new.append(org.get(i, ''))
            elif i in data:
                new.append(data[i])
            else:
                new.append(org.get(i, ''))
        
        return new

    
    def GetCol(self, title:str):
        title = title.lower()
        if not title in self.head:return
        return self.head.index(title) + 1
    
    def GetHead(self):
        head = []
        for uid in self.ws[1]:
            head.append(uid.value)
        return head
    
    def SearchCert(self, key, col):
        cty = type(col)
        if cty != int:
            if cty == str:
                col = self.GetCol(col)
            if cty in [list, tuple]:
                return
        
        sec = []
        for ron in range(2, self.ws.max_row + 1):
            val = self.ws.cell(row=ron, column=col).value
            if val == key:sec.append(ron)
        return sec
    
    def Append(self, data:list):
        self.ws.append(data)
        self.Save()
    
    def GetRows(self, data:list):
        a = []
        for num in data:
            if num > self.ws.max_row + 1:continue
            n = []
            for uid in self.ws[num]:
                n.append(uid.value)
            a.append( dict(zip(self.head, n)) )
        return a
    
    def WriteRow(self, row:int, data:list):
        total = len(self.ws[row])
        if len(data) < total:return

        for num in range(0, total):
            blk = self.ws[row][num]
            blk.value = data[num]
        self.Save()
    
    def DeleteRows(self, data:list):
        for i in data:
            try:self.ws.delete_rows(i)
            except:pass
        self.Save()
    
    def Save(self):
        self.wb.save(self.location)